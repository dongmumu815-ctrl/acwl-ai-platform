"""Excel生成服务"""

import io
import json
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    import re
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from minio import Minio
from minio.error import S3Error

from app.core.config import settings
from app.models.resource_package import ResourcePackage, ResourcePackageFile
from app.services.elasticsearch_service import ElasticsearchService
import logging

logger = logging.getLogger(__name__)


class ExcelService:
    """Excel生成和MinIO上传服务"""
    
    def __init__(self):
        self.minio_client = Minio(
            endpoint=settings.MINIO_ENDPOINT,
            access_key=settings.MINIO_ACCESS_KEY,
            secret_key=settings.MINIO_SECRET_KEY,
            secure=settings.MINIO_SECURE,
            region=settings.MINIO_REGION
        )
        self.bucket_name = settings.MINIO_BUCKET_NAME
        self.es_service = ElasticsearchService()
        self._ensure_bucket_exists()
    
    def _ensure_bucket_exists(self):
        """确保存储桶存在"""
        try:
            if not self.minio_client.bucket_exists(self.bucket_name):
                self.minio_client.make_bucket(self.bucket_name, location=settings.MINIO_REGION)
                logger.info(f"创建MinIO存储桶: {self.bucket_name}")
        except S3Error as e:
            logger.error(f"MinIO存储桶操作失败: {e}")
            raise
    
    async def generate_and_upload_excel(
        self, 
        package_id: int, 
        query_data: Dict[str, Any],
        db: AsyncSession
    ) -> Dict[str, Any]:
        """
        生成Excel并上传到MinIO
        
        Args:
            package_id: 资源包ID
            query_data: 查询数据和参数
            db: 数据库会话
            
        Returns:
            Dict包含下载URL和状态信息
        """
        # 获取资源包信息
        result = await db.execute(
            select(ResourcePackage).where(ResourcePackage.id == package_id)
        )
        package = result.scalar_one_or_none()
        
        if not package:
            raise ValueError("资源包不存在")
        
        # 执行查询获取最新数据（基于 excel_time 过滤）
        query_result = await self._execute_query(package, query_data, db)
        
        if not query_result or not query_result.get('hits', {}).get('hits'):
            return {"status": "no_new_data", "message": "无最新数据，无需生成"}
        
        # 检查数据创建时间是否大于 excel_time
        applied_time_filter = package.excel_time is not None
        if not await self._has_new_data(package, query_result, applied_time_filter):
            return {"status": "no_new_data", "message": "无最新数据，无需生成"}
        
        # 生成Excel文件
        excel_buffer = self._generate_excel(query_result, package.name, requested_fields=query_data.get("_source"))
        
        # 上传到MinIO
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"resource_package_{package_id}_{timestamp}.xlsx"
        object_path = f"resource_packages/{package_id}/{filename}"
        
        try:
            self.minio_client.put_object(
                bucket_name=self.bucket_name,
                object_name=object_path,
                data=excel_buffer,
                length=excel_buffer.getbuffer().nbytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # 保存历史记录
            file_rec = ResourcePackageFile(
                package_id=package_id,
                filename=filename,
                object_path=object_path,
                generated_at=datetime.now()
            )
            db.add(file_rec)
            
            # 更新资源包的生成信息
            package.excel_time = datetime.now()
            package.download_url = f"minio://{settings.MINIO_ENDPOINT}/{self.bucket_name}/{object_path}"
            await db.commit()
            
            # 生成预签名下载URL（有效期1小时）
            download_url = self.minio_client.presigned_get_object(
                bucket_name=self.bucket_name,
                object_name=object_path,
                expires=timedelta(seconds=3600)  # 1小时
            )
            
            return {
                "status": "success",
                "download_url": download_url,
                "filename": filename,
                "minio_path": package.download_url
            }
            
        except S3Error as e:
            logger.error(f"MinIO上传失败: {e}")
            raise Exception(f"文件上传失败: {str(e)}")
    
    async def _execute_query(self, package: ResourcePackage, query_data: Dict[str, Any], db: AsyncSession) -> Dict[str, Any]:
        """执行ES查询获取数据（按 excel_time 过滤）"""
        try:
            # 获取数据源信息
            from app.services.datasource import DatasourceService
            from elasticsearch import AsyncElasticsearch
            from elasticsearch.exceptions import ConnectionError as ESConnectionError, AuthenticationException, ConnectionTimeout as ESConnectionTimeout
            try:
                # ES 8+ 客户端的通用 API 错误类型
                from elasticsearch import ApiError
            except Exception:
                # 兼容旧版本
                from elasticsearch.exceptions import TransportError as ApiError
            
            datasource_service = DatasourceService(db)
            datasource = await datasource_service.get_datasource(package.datasource_id)
            
            if not datasource:
                raise ValueError(f"数据源 ID {package.datasource_id} 不存在")
            
            # 连接参数（超时/重试/协议）
            timeout = datasource.connection_params.get('timeout', 30) if datasource.connection_params else 30
            # 归一化与最小阈值：避免过低的超时导致频繁失败
            try:
                timeout = int(timeout)
            except Exception:
                timeout = 30
            if timeout < 30:
                logger.info(f"检测到ES超时配置过低({timeout}s)，使用最小超时30s")
                timeout = 30

            max_retries = datasource.connection_params.get('max_retries', 3) if datasource.connection_params else 3
            try:
                max_retries = int(max_retries)
            except Exception:
                max_retries = 3
            scheme = 'http'
            if datasource.connection_params:
                params = datasource.connection_params
                if isinstance(params.get('use_ssl'), bool) and params.get('use_ssl'):
                    scheme = 'https'
                elif params.get('scheme') in ('http', 'https'):
                    scheme = params.get('scheme')
            logger.info(
                f"准备执行ES查询: datasource_id={package.datasource_id}, host={datasource.host}, port={datasource.port}, scheme={scheme}, "
                f"timeout={timeout}, max_retries={max_retries}, index={query_data.get('index')}"
            )

            # 创建ES客户端
            if datasource.username and datasource.password:
                es_client = AsyncElasticsearch(
                    [{
                        'scheme': scheme,
                        'host': datasource.host,
                        'port': datasource.port
                    }],
                    basic_auth=(datasource.username, datasource.password),
                    verify_certs=False,
                    request_timeout=timeout,
                    max_retries=max_retries,
                    retry_on_timeout=True
                )
            else:
                es_client = AsyncElasticsearch(
                    [{
                        'scheme': scheme,
                        'host': datasource.host,
                        'port': datasource.port
                    }],
                    verify_certs=False,
                    request_timeout=timeout,
                    max_retries=max_retries,
                    retry_on_timeout=True
                )
            
            try:
                # 预检查：集群健康，便于快速定位连接/认证问题
                try:
                    await es_client.cluster.health(request_timeout=timeout)
                    logger.info("ES健康检查通过")
                except ESConnectionTimeout as e:
                    logger.error(f"ES健康检查超时 [host={datasource.host} port={datasource.port} scheme={scheme} timeout={timeout}] : {e}")
                    raise
                except ESConnectionError as e:
                    logger.error(f"ES连接失败 [host={datasource.host} port={datasource.port} scheme={scheme}] : {e}")
                    raise
                except AuthenticationException as e:
                    logger.error(f"ES认证失败 [host={datasource.host} port={datasource.port} scheme={scheme}] : {e}")
                    raise

                # 构建查询体（加入 create_time > excel_time 过滤）
                base_query = query_data.get("query", {"match_all": {}})
                
                # 根据资源包的excel生成时间添加 create_time 过滤
                xl_time = package.excel_time
                range_filter = None
                if xl_time is not None:
                    xl_dt = None
                    if isinstance(xl_time, (int, float)):
                        try:
                            xl_dt = datetime.fromtimestamp(xl_time)
                        except Exception:
                            xl_dt = None
                    elif isinstance(xl_time, datetime):
                        xl_dt = xl_time
                    if xl_dt is not None:
                        excel_time_str = xl_dt.strftime("%Y-%m-%dT%H:%M:%S")
                        range_filter = {"range": {"create_time": {"gt": excel_time_str}}}
                
                # 合成最终查询
                if range_filter:
                    query = {"bool": {"must": [base_query], "filter": [range_filter]}}
                else:
                    query = base_query
                
                # 滚动查询参数：单页大小与总上限（默认5万，可通过数据源配置覆盖）
                params = datasource.connection_params or {}
                try:
                    export_max_rows = int(params.get('export_max_rows', 50000))
                except Exception:
                    export_max_rows = 50000
                try:
                    page_size = int(params.get('export_page_size', 5000))
                except Exception:
                    page_size = 5000

                # 保留用户选择的字段，同时强制包含创建时间字段，避免_source裁剪导致误判
                source_fields = None
                if query_data.get("_source"):
                    src = query_data["_source"]
                    if isinstance(src, list):
                        ensure_fields = ["create_time", "created_at", "createTime"]
                        source_fields = list({*src, *ensure_fields})
                    else:
                        source_fields = src

                # 执行滚动查询：按 _doc 排序以获得稳定的分页性能
                all_hits = []
                total_value = None
                took_total = 0
                pages_fetched = 0
                scroll_id = None
                scroll_keep_alive = "2m"

                try:
                    first = await es_client.search(
                        index=query_data.get("index", ""),
                        query=query,
                        size=page_size,
                        sort=["_doc"],
                        source=source_fields,
                        scroll=scroll_keep_alive,
                        request_timeout=timeout,
                        allow_partial_search_results=True
                    )
                except ESConnectionTimeout as e:
                    logger.error(f"ES查询超时 [index={query_data.get('index')} host={datasource.host} port={datasource.port} scheme={scheme} timeout={timeout}] : {e}")
                    raise
                except ApiError as e:
                    # 详细记录错误信息，并进行一次回退：去掉sort重试
                    detail = getattr(e, 'body', None) or str(e)
                    logger.error(f"ES初始滚动查询失败(ApiError): {detail}")
                    logger.info("尝试回退：移除sort后再次执行初始查询")
                    first = await es_client.search(
                        index=query_data.get("index", ""),
                        query=query,
                        size=page_size,
                        source=source_fields,
                        scroll=scroll_keep_alive,
                        request_timeout=timeout,
                        allow_partial_search_results=True
                    )

                try:
                    total_meta = first.get('hits', {}).get('total')
                    total_value = (total_meta.get('value') if isinstance(total_meta, dict) else total_meta) if total_meta is not None else None
                    took_total += int(first.get('took', 0) or 0)
                    batch_hits = first.get('hits', {}).get('hits', []) or []
                    all_hits.extend(batch_hits)
                    pages_fetched += 1
                    scroll_id = first.get('_scroll_id')
                    logger.info(f"ES滚动查询初始页: hits_count={len(batch_hits)}, total={total_value}, took={first.get('took')}ms, page_size={page_size}, cap={export_max_rows}")
                except Exception:
                    logger.info("ES滚动查询初始页: 无法解析摘要信息")

                # 继续滚动直到达到上限或无更多数据
                try:
                    while scroll_id and len(all_hits) < export_max_rows:
                        try:
                            next_page = await es_client.scroll(scroll_id=scroll_id, scroll=scroll_keep_alive, request_timeout=timeout)
                        except ApiError as e:
                            detail = getattr(e, 'body', None) or str(e)
                            logger.error(f"ES滚动查询下一页失败(ApiError): {detail}")
                            break
                        batch_hits = next_page.get('hits', {}).get('hits', []) or []
                        if not batch_hits:
                            logger.info("ES滚动查询：无更多数据，结束滚动")
                            break
                        all_hits.extend(batch_hits)
                        pages_fetched += 1
                        took_total += int(next_page.get('took', 0) or 0)
                        scroll_id = next_page.get('_scroll_id') or scroll_id
                        logger.info(f"ES滚动查询第{pages_fetched}页: 累计={len(all_hits)} / cap={export_max_rows}")
                finally:
                    if scroll_id:
                        try:
                            await es_client.clear_scroll(scroll_id=scroll_id)
                            logger.info("已清理ES滚动上下文")
                        except Exception as ce:
                            logger.warning(f"清理ES滚动上下文失败: {ce}")

                # 截断到上限
                if len(all_hits) > export_max_rows:
                    all_hits = all_hits[:export_max_rows]

                # 汇总返回结构与摘要
                result = {
                    'hits': {
                        'total': total_value,
                        'hits': all_hits
                    },
                    'took': took_total,
                    'scroll_used': True,
                    'pages_fetched': pages_fetched,
                    'page_size': page_size,
                    'export_max_rows': export_max_rows
                }
                logger.info(f"ES滚动查询完成: total={total_value}, returned={len(all_hits)}, pages={pages_fetched}, took_total={took_total}ms")
                return result
                
            finally:
                await es_client.close()
            
        except Exception as e:
            logger.error(
                f"执行ES查询失败: {e} | datasource_id={package.datasource_id}, host={getattr(datasource, 'host', 'unknown')}, "
                f"port={getattr(datasource, 'port', 'unknown')}, index={query_data.get('index')}"
            )
            raise
    
    async def _has_new_data(self, package: ResourcePackage, query_result: Dict[str, Any], applied_time_filter: bool = False) -> bool:
        """检查是否有新数据（与 excel_time 比较）

        applied_time_filter: 查询阶段是否已应用了基于 excel_time 的时间过滤
        """
        hits = query_result.get('hits', {}).get('hits', [])
        if not hits:
            return False
        
        # 如果是第一次生成（excel_time为None），有数据即可生成
        if package.excel_time is None:
            return True
        
        # 确保excel_time是datetime对象
        excel_time = package.excel_time
        if isinstance(excel_time, (int, float)):
            try:
                # 检查时间戳是否有效（不能为负数或过大）
                if excel_time < 0 or excel_time > 2147483647:  # Unix时间戳最大值
                    logger.warning(f"无效的时间戳值: {excel_time}")
                    return True
                excel_time = datetime.fromtimestamp(excel_time)
            except (ValueError, OSError, OverflowError) as e:
                logger.warning(f"时间戳转换失败: {e}, excel_time: {excel_time}")
                return True  # 转换失败，默认认为有新数据
        elif not isinstance(excel_time, datetime):
            logger.warning(f"excel_time类型异常: {type(excel_time)}, 值: {excel_time}")
            return True  # 如果类型异常，默认认为有新数据
        
        # 检查数据中是否有create_time字段大于excel_time的记录
        for hit in hits:
            source = hit.get('_source', {})
            create_time_str = source.get('create_time') or source.get('created_at') or source.get('createTime')
            
            if create_time_str:
                try:
                    # 尝试解析时间字符串
                    if isinstance(create_time_str, str):
                        # 支持多种时间格式
                        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"]:
                            try:
                                create_time = datetime.strptime(create_time_str, fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            continue  # 无法解析时间格式，跳过此记录
                    elif isinstance(create_time_str, (int, float)):
                        # 时间戳格式
                        try:
                            # 检查时间戳是否有效
                            if create_time_str < 0 or create_time_str > 2147483647:
                                logger.warning(f"无效的create_time时间戳: {create_time_str}")
                                continue
                            create_time = datetime.fromtimestamp(create_time_str)
                        except (ValueError, OSError, OverflowError) as e:
                            logger.warning(f"create_time时间戳转换失败: {e}, create_time_str: {create_time_str}")
                            continue
                    else:
                        continue
                    
                    # 比较时间
                    if create_time > excel_time:
                        return True
                        
                except (ValueError, TypeError) as e:
                    logger.warning(f"时间解析失败: {e}, create_time_str: {create_time_str}")
                    continue
        
        # 兜底：如果查询阶段已应用了时间过滤且存在命中，则视为有新数据
        if applied_time_filter and hits:
            return True
        
        return False
    
    def _generate_excel(self, query_result: Dict[str, Any], package_name: str, requested_fields: Optional[List[str]] = None) -> io.BytesIO:
        """生成Excel文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "查询结果"
        
        # 获取数据
        hits = query_result.get('hits', {}).get('hits', [])
        if not hits:
            raise ValueError("无数据可导出")
        
        # 获取所有字段名
        all_fields = set()
        for hit in hits:
            source = hit.get('_source', {})
            all_fields.update(source.keys())
        
        # 优先使用请求中的_source字段列表，以保持与前端显示一致
        fields = list(requested_fields) if requested_fields and len(requested_fields) > 0 else sorted(list(all_fields))
        
        # 设置标题样式
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
        
        # 写入数据
        for row_idx, hit in enumerate(hits, 2):
            source = hit.get('_source', {})
            for col_idx, field in enumerate(fields, 1):
                value = source.get(field, '')
                # 处理复杂数据类型
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                # 去除非法控制字符，避免 openpyxl IllegalCharacterError
                if isinstance(value, str):
                    value = ILLEGAL_CHARACTERS_RE.sub('', value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 自动调整列宽
        for col_idx, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = min(max(len(field) + 2, 10), 50)
        
        # 添加信息工作表
        info_ws = wb.create_sheet("导出信息")
        # 清理包名中的非法字符用于信息工作表
        safe_package_name = ILLEGAL_CHARACTERS_RE.sub('', package_name) if isinstance(package_name, str) else package_name
        info_data = [
            ["资源包名称", safe_package_name],
            ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["数据条数", len(hits)],
            ["字段数量", len(fields)]
        ]
        
        for row_idx, (key, value) in enumerate(info_data, 1):
            info_ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            info_ws.cell(row=row_idx, column=2, value=value)
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer